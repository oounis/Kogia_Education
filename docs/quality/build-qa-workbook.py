#!/usr/bin/env python3
"""Generate the Coreon EDU manual QA workbook template."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------- palette
NAVY   = "1F2A44"
PURPLE = "7539E4"
LIGHT  = "F1F3F8"
BORDER = "D5DAE5"
GREEN  = "C6EFCE"; GREEN_T  = "0B6B2F"
RED    = "FFC7CE"; RED_T    = "9C0006"
AMBER  = "FFE699"; AMBER_T  = "8A5A00"
GREY   = "E4E7EE"; GREY_T   = "50596B"

thin = Side(style="thin", color=BORDER)
BOX  = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()


def sheet(title, headers, widths, notes=None, freeze="A3", tab=PURPLE):
    """Create a sheet: title bar + header row + styling + autofilter."""
    ws = wb.create_sheet(title)
    ws.sheet_properties.tabColor = tab
    n = len(headers)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    t = ws.cell(row=1, column=1, value=notes or title)
    t.font = Font(bold=True, size=12, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=NAVY)
    t.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[1].height = 26

    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = Font(bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=PURPLE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BOX
        ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
    ws.row_dimensions[2].height = 42

    ws.freeze_panes = freeze
    ws.auto_filter.ref = f"A2:{get_column_letter(n)}2"
    return ws


def dropdown(ws, formula, col, first=3, last=1000):
    dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(f"{col}{first}:{col}{last}")
    return dv


def colour_rule(ws, col, mapping, first=3, last=1000):
    for value, (bg, fg) in mapping.items():
        ws.conditional_formatting.add(
            f"{col}{first}:{col}{last}",
            CellIsRule(operator="equal", formula=[f'"{value}"'],
                       fill=PatternFill("solid", bgColor=bg),
                       font=Font(color=fg, bold=True)),
        )


def body_style(ws, ncols, first=3, last=400):
    for r in range(first, last + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BOX
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(size=10)


# ================================================================ 1. README
ws = wb.active
ws.title = "READ ME"
ws.sheet_properties.tabColor = NAVY
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 108

readme = [
    ("H1", "COREON EDU — MANUAL TEST EXECUTION WORKBOOK", ""),
    ("P",  "Standard", "Test Case Specification + Defect Register, per ISO/IEC/IEEE 29119-3 (formerly IEEE 829)."),
    ("P",  "Phase", "Test Design & Test Execution, inside the System Testing stage of the STLC."),
    ("P",  "Product", "Coreon EDU — Kogia Group"),
    ("P",  "Owner", "Othman Ounis"),
    ("", "", ""),
    ("H2", "THE SHEETS", ""),
    ("R", "Dashboard", "Read-only. Auto-calculates pass rate, coverage and open defects. Do not type here."),
    ("R", "Test Log", "THE MAIN SHEET. One row per element tested. This is where you spend your time."),
    ("R", "Defect Log", "One row per BUG — something that is broken in what already exists. Every FAIL in Test Log points here."),
    ("R", "Change Requests", "One row per thing to ADD or CHANGE. Not bugs. Enhancements, new features, design, content, epics, research."),
    ("R", "Raw Notes (original)", "Your notes, kept word for word. Nothing is ever deleted from here — it is the audit trail."),
    ("R", "Modules", "Inventory of all 46 screens. Tick them off as you cover them. Gives you the coverage %."),
    ("R", "Roles & Logins", "The 7 demo accounts. Use these, never invent new ones — results must be reproducible."),
    ("R", "Test Cycles", "One row per round of testing (per release). Records the Go / No-Go decision."),
    ("R", "Lists", "Source data for the dropdowns. Edit only if you want to change the allowed values."),
    ("", "", ""),
    ("H2", "HOW TO USE IT — THE LOOP", ""),
    ("N", "1. Pick a module", "Take one row from the Modules sheet. Test it completely before moving on. Do not jump around."),
    ("N", "2. Pick a role", "Test the module as ONE role at a time. The same button can behave differently for a parent and for Direction."),
    ("N", "3. Click everything", "Every button, link, tab, dropdown, toggle, icon, table header, modal close. One row in Test Log per element."),
    ("N", "4. Write Expected BEFORE Actual", "Decide what SHOULD happen before you click. If you click first, you will accept whatever you see as normal. This is the single most important discipline in this file."),
    ("N", "5. Mark the status", "PASS / FAIL / BLOCKED / NOT RUN / N-A. Blocked means you could not test it because something else is broken — it is not a Fail."),
    ("N", "6. Log the defect", "Every FAIL gets a row in Defect Log and its ID written back into the Test Log row."),
    ("N", "7. Retest after the fix", "Set Retest to PASS or FAIL. Never close a defect without retesting it yourself."),
    ("", "", ""),
    ("H2", "DEFECT vs CHANGE REQUEST — THE MOST IMPORTANT DISTINCTION IN THIS FILE", ""),
    ("N", "A DEFECT is", "The product does not do what it was supposed to do. Something is broken. It goes in the Defect Log."),
    ("N", "A CHANGE REQUEST is", "The product works as designed, but the design itself should change. Nothing is broken. It goes in Change Requests."),
    ("N", "Why keep them apart", "They are decided by different people, cost different money, and answer different questions. A defect asks 'when is it fixed'. A change request asks 'do we even want it, and is it worth the price'. Mixing them into one list is the fastest way to lose control of a product."),
    ("N", "The test", "Ask: was it ever meant to work this way? If yes and it does not -> defect. If it works as intended and you simply want something else -> change request."),
    ("N", "Example - defect", "The login page does not fit the browser window. It was never meant to overflow. DEF-001."),
    ("N", "Example - change request", "Make the weather widget smaller. It works exactly as built; you want a different design. CR-015."),
    ("", "", ""),
    ("H2", "SEVERITY vs PRIORITY — THEY ARE NOT THE SAME", ""),
    ("N", "Severity", "How much damage the bug does. Set by the tester. Critical / Major / Minor / Trivial."),
    ("N", "Priority", "How fast it must be fixed. Set by the product owner. P1 / P2 / P3 / P4."),
    ("N", "Example", "A typo on the landing page is Trivial severity — but P1 priority if a school is visiting the site tomorrow."),
    ("", "", ""),
    ("H2", "SEVERITY SCALE", ""),
    ("N", "Critical", "Data loss, data leak between roles, money is wrong, or the app cannot be used at all. Stop the release."),
    ("N", "Major", "A main function does not work, and there is no workaround."),
    ("N", "Minor", "A function works but behaves wrongly, or there is an acceptable workaround."),
    ("N", "Trivial", "Cosmetic: spacing, wording, colour, alignment."),
    ("", "", ""),
    ("H2", "RULES THAT MAKE THIS WORKBOOK WORTH KEEPING", ""),
    ("N", "Be reproducible", "Anyone must be able to read a row and reproduce your result exactly. If they cannot, the row is worthless."),
    ("N", "One element, one row", "Never write 'tested the whole page'. That hides everything you did not check."),
    ("N", "Never delete a row", "A test that no longer applies is marked N-A, not erased. History is the point."),
    ("N", "Test as the user, not the author", "Write your Expected Result from the specification, not from your own code. Otherwise you only test what you remembered to build."),
    ("N", "Record the build", "A result without a build number cannot be trusted a week later."),
    ("", "", ""),
    ("H2", "WHAT AUTOMATION ALREADY COVERS — DO NOT REDO IT BY HAND", ""),
    ("N", "Automated suites", "56 core tests, 11 server tests, 17 Playwright end-to-end journeys, all gating the deploy in CI."),
    ("N", "Dead-button audit", "e2e/outil.audit-clics.mjs already clicks every button on every page for 3 roles and reports dead ones."),
    ("N", "Spend your manual time on", "Does it look right · Does it read right in Arabic (RTL) · Is the business rule correct · Is the number correct · Does it make sense to a real school."),
]

r = 1
for kind, a, b in readme:
    if kind == "H1":
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        c = ws.cell(row=r, column=1, value=a)
        c.font = Font(bold=True, size=16, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[r].height = 34
    elif kind == "H2":
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        c = ws.cell(row=r, column=1, value=a)
        c.font = Font(bold=True, size=11, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=PURPLE)
        c.alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[r].height = 24
    elif kind in ("P", "R", "N"):
        ca = ws.cell(row=r, column=1, value=a)
        ca.font = Font(bold=True, size=10)
        ca.alignment = Alignment(vertical="top", indent=1)
        if kind == "R":
            ca.fill = PatternFill("solid", fgColor=LIGHT)
        cb = ws.cell(row=r, column=2, value=b)
        cb.font = Font(size=10)
        cb.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = 30 if len(b) > 95 else 16
    r += 1


# ================================================================ 2. LISTS
lists = wb.create_sheet("Lists")
lists.sheet_properties.tabColor = GREY_T

LIST_DATA = {
    "Status":        ["NOT RUN", "PASS", "FAIL", "BLOCKED", "N-A"],
    "Severity":      ["Critical", "Major", "Minor", "Trivial"],
    "Priority":      ["P1 - Immediate", "P2 - High", "P3 - Medium", "P4 - Low"],
    "ElementType":   ["Page load", "Button", "Link", "Form field", "Form submit", "Dropdown",
                      "Checkbox", "Radio", "Toggle", "Tab", "Modal", "Table", "Table sort",
                      "Search", "Filter chip", "Stat tile", "Icon", "Menu", "Drag & drop",
                      "File upload", "Export CSV", "Print / PDF", "Keyboard shortcut",
                      "Notification", "Toast", "Navigation", "Empty state", "Error state"],
    "TestType":      ["Functional", "UI / Visual", "Negative", "Access control", "Business rule",
                      "Data integrity", "Responsive", "RTL / Arabic", "Accessibility",
                      "Performance", "Security", "Regression", "Exploratory"],
    "Role":          ["owner", "schooladmin", "admin", "teacher", "supervisor", "security",
                      "parent", "not logged in"],
    "Environment":   ["Local dev", "Local build", "Server mode", "GitHub Pages", "edu.kogiagroup.com"],
    "Device":        ["Desktop Chrome", "Desktop Firefox", "Desktop Safari", "Desktop Edge",
                      "Mobile Chrome 390px", "Mobile Safari iOS", "Tablet", "Expo mobile app"],
    "DefectStatus":  ["NEW", "OPEN", "IN PROGRESS", "FIXED", "READY TO RETEST", "CLOSED",
                      "REOPENED", "WON'T FIX", "DUPLICATE", "NOT A BUG"],
    "Repro":         ["Always", "Sometimes", "Once only", "Not reproduced"],
    "Retest":        ["Not retested", "PASS", "FAIL"],
    "ModuleStatus":  ["Not started", "In progress", "Complete", "Blocked", "Out of scope"],
    "YesNo":         ["Yes", "No"],
    "Decision":      ["GO", "NO-GO", "GO with conditions", "Pending"],
    "CRType":        ["Enhancement", "New feature", "Design", "Content", "Epic", "Research"],
    "CRStatus":      ["NEW", "TRIAGED", "APPROVED", "IN PROGRESS", "DONE", "REJECTED", "DEFERRED"],
    "Effort":        ["XS", "S", "M", "L", "XL"],
}

for i, (name, values) in enumerate(LIST_DATA.items(), start=1):
    col = get_column_letter(i)
    h = lists.cell(row=1, column=i, value=name)
    h.font = Font(bold=True, size=10, color="FFFFFF")
    h.fill = PatternFill("solid", fgColor=NAVY)
    h.alignment = Alignment(horizontal="center")
    lists.column_dimensions[col].width = max(14, len(name) + 6)
    for j, v in enumerate(values, start=2):
        lists.cell(row=j, column=i, value=v).font = Font(size=10)

def named(key):
    i = list(LIST_DATA).index(key) + 1
    col = get_column_letter(i)
    return f"=Lists!${col}$2:${col}${len(LIST_DATA[key]) + 1}"

lists.freeze_panes = "A2"


# ================================================================ 3. TEST LOG
TL_HEADERS = [
    "Test ID", "Date", "Tester", "Build / Commit", "Environment", "Browser / Device",
    "Module / Feature", "Sub-module", "Screen / Page", "Route (URL)",
    "Element type", "Element name (label on screen)", "Role under test",
    "Test title", "Preconditions", "Test steps", "Test data",
    "Expected result", "Actual result",
    "Status", "Severity", "Priority", "Defect ID", "Evidence (screenshot / file)",
    "Retest", "Retest date", "Comments",
]
TL_WIDTHS = [10, 11, 12, 15, 14, 17, 18, 16, 18, 22, 15, 30, 14, 34, 26, 40, 20,
             38, 38, 11, 11, 15, 11, 24, 13, 11, 26]

tl = sheet("Test Log", TL_HEADERS, TL_WIDTHS,
           notes="TEST LOG — one row per element tested.  Fill Expected result BEFORE you click.  "
                 "Every FAIL needs a Defect ID.",
           freeze="G3")
body_style(tl, len(TL_HEADERS), last=600)

dropdown(tl, named("Environment"), "E", last=600)
dropdown(tl, named("Device"), "F", last=600)
dropdown(tl, named("ElementType"), "K", last=600)
dropdown(tl, named("Role"), "M", last=600)
dropdown(tl, named("Status"), "T", last=600)
dropdown(tl, named("Severity"), "U", last=600)
dropdown(tl, named("Priority"), "V", last=600)
dropdown(tl, named("Retest"), "Y", last=600)

colour_rule(tl, "T", {"PASS": (GREEN, GREEN_T), "FAIL": (RED, RED_T),
                      "BLOCKED": (AMBER, AMBER_T), "NOT RUN": (GREY, GREY_T),
                      "N-A": (GREY, GREY_T)}, last=600)
colour_rule(tl, "U", {"Critical": (RED, RED_T), "Major": (AMBER, AMBER_T)}, last=600)
colour_rule(tl, "Y", {"PASS": (GREEN, GREEN_T), "FAIL": (RED, RED_T)}, last=600)

# two worked examples so the format is unambiguous
EX = [
    ["TC-001", "=TODAY()", "Othman", "50d4c48", "Local dev", "Desktop Chrome",
     "Authentication", "Login", "Login", "#/login", "Form submit",
     "Se connecter (submit button)", "not logged in",
     "Login is refused with a wrong password",
     "App is open at #/login. Demo data seeded.",
     "1. Type parent@alnour.tn in E-mail\n2. Type wrongpass in Mot de passe\n3. Click Se connecter",
     "parent@alnour.tn / wrongpass",
     "Error message 'E-mail ou mot de passe incorrect.' is shown. User stays on #/login. No session created.",
     "", "NOT RUN", "", "", "", "", "Not retested", "", "Example row — delete or overwrite"],
    ["TC-002", "=TODAY()", "Othman", "50d4c48", "Local dev", "Mobile Chrome 390px",
     "Évaluation", "Save step", "Evaluate", "#/app/evaluate", "Button",
     "Enregistrer", "teacher",
     "Saving with zero students placed is blocked",
     "Logged in as enseignant@alnour.tn. A slot is selected. No student dragged into any bucket.",
     "1. Open #/app/evaluate\n2. Select the live slot\n3. Place no student\n4. Click Enregistrer",
     "Class CP-A, no placements",
     "An error toast appears, nothing is saved, no parent notification is sent, the button stays enabled.",
     "", "NOT RUN", "", "", "", "", "Not retested", "", "Example row — delete or overwrite"],
]
for i, row in enumerate(EX, start=3):
    for j, v in enumerate(row, start=1):
        c = tl.cell(row=i, column=j, value=v)
        c.font = Font(size=10, italic=True, color="7A8194")
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = BOX
    tl.row_dimensions[i].height = 76

# ID auto-suggestion is left manual on purpose (formulas break on sort)


# ================================================================ 4. DEFECT LOG
DL_HEADERS = [
    "Defect ID", "Date found", "Found by", "Linked Test ID", "Build / Commit",
    "Module / Feature", "Screen / Page", "Route (URL)", "Element", "Role",
    "Summary (one line)", "Steps to reproduce", "Expected result", "Actual result",
    "Severity", "Priority", "Reproducibility", "Environment", "Browser / Device",
    "Evidence", "Status", "Assigned to", "Fixed in build", "Date closed",
    "Root cause", "Comments",
]
DL_WIDTHS = [11, 12, 12, 12, 15, 18, 18, 22, 24, 13, 40, 40, 34, 34,
             11, 15, 15, 14, 17, 22, 16, 13, 14, 12, 28, 26]

dl = sheet("Defect Log", DL_HEADERS, DL_WIDTHS,
           notes="DEFECT LOG — one row per bug.  Severity = how bad the damage (tester decides).  "
                 "Priority = how fast to fix (owner decides).  They are not the same.",
           freeze="F3", tab=RED_T)
body_style(dl, len(DL_HEADERS), last=400)

dropdown(dl, named("Severity"), "O", last=400)
dropdown(dl, named("Priority"), "P", last=400)
dropdown(dl, named("Repro"), "Q", last=400)
dropdown(dl, named("Environment"), "R", last=400)
dropdown(dl, named("Device"), "S", last=400)
dropdown(dl, named("DefectStatus"), "U", last=400)
dropdown(dl, named("Role"), "J", last=400)

colour_rule(dl, "O", {"Critical": (RED, RED_T), "Major": (AMBER, AMBER_T),
                      "Minor": (GREY, GREY_T), "Trivial": (GREY, GREY_T)}, last=400)
colour_rule(dl, "U", {"CLOSED": (GREEN, GREEN_T), "FIXED": (GREEN, GREEN_T),
                      "NEW": (RED, RED_T), "OPEN": (RED, RED_T),
                      "REOPENED": (RED, RED_T),
                      "IN PROGRESS": (AMBER, AMBER_T),
                      "READY TO RETEST": (AMBER, AMBER_T)}, last=400)

ex_bug = ["BUG-001", "=TODAY()", "Othman", "TC-001", "50d4c48",
          "Authentication", "Setup wizard", "#/setup", "Whole page", "not logged in",
          "The setup wizard is reachable without logging in",
          "1. Open a private window\n2. Go to #/setup directly\n3. The wizard renders",
          "Redirect to #/login — the wizard is Direction-only.",
          "The wizard opens and school settings can be overwritten.",
          "Critical", "P1 - Immediate", "Always", "Local dev", "Desktop Chrome",
          "setup-no-auth.png", "NEW", "Othman", "", "", "Route sits outside the Protected guard",
          "Example row — delete or overwrite"]
for j, v in enumerate(ex_bug, start=1):
    c = dl.cell(row=3, column=j, value=v)
    c.font = Font(size=10, italic=True, color="7A8194")
    c.alignment = Alignment(vertical="top", wrap_text=True)
    c.border = BOX
dl.row_dimensions[3].height = 76


# ================================================================ 5. MODULES
MODULES = [
    # (module, route, roles, flag/notes)
    ("Landing page", "#/", "public", "Marketing, 8 CTAs, FAQ accordion"),
    ("Login", "#/login", "public", "7 demo role buttons, FR/AR switch"),
    ("Setup wizard", "#/setup", "schooladmin", "First-run; currently outside the auth guard"),
    ("Pre-enrolment (public)", "#/inscription", "public", "Public form + attachments"),
    ("Dashboard", "#/app", "all 7", "6 different dashboards in one page"),
    ("Schools (SaaS console)", "#/app/schools", "owner", ""),
    ("Settings", "#/app/settings", "schooladmin", "6 tabs"),
    ("Accounts", "#/app/accounts", "schooladmin", "Last Direction account is protected"),
    ("Students", "#/app/students", "schooladmin, admin, supervisor, teacher", "DataTable + 24-field modal"),
    ("Student profile (360)", "#/app/eleve/:id", "schooladmin, admin, supervisor, teacher", ""),
    ("Teachers", "#/app/teachers", "schooladmin, admin", "14-field add modal"),
    ("Staff attendance", "#/app/staff", "schooladmin, admin", "3 tabs, CSV export"),
    ("HR", "#/app/hr", "schooladmin, admin", "Payroll: prepare > recalc > validate > paid"),
    ("Accounting", "#/app/accounting", "schooladmin, admin", "Invoices immutable, credit notes"),
    ("Academic", "#/app/academic", "schooladmin, admin, teacher", "Year promotion is irreversible"),
    ("Facilities", "#/app/facilities", "schooladmin, admin", "Booking conflicts"),
    ("Pointage (clock-in)", "#/app/pointage", "teacher, supervisor, security, admin", ""),
    ("Evaluate", "#/app/evaluate", "teacher", "FLAGSHIP FLOW — drag & drop, draft autosave"),
    ("Child file", "#/app/child", "teacher, admin, schooladmin", "Pickup authorisations, health"),
    ("Day journal", "#/app/journal", "teacher, admin, schooladmin, parent", "Parent is read-only"),
    ("Results", "#/app/results", "schooladmin, admin", ""),
    ("Timetable", "#/app/timetable", "schooladmin, admin, teacher, parent, supervisor", ""),
    ("Attendance", "#/app/attendance", "schooladmin, teacher, admin, supervisor", "Roll call + notify parent"),
    ("Homework", "#/app/homework", "teacher, admin, parent", "FEATURE FLAG OFF by default"),
    ("Exams", "#/app/exams", "schooladmin, admin, teacher, parent", "FEATURE FLAG OFF by default"),
    ("Finance", "#/app/finance", "schooladmin, admin", ""),
    ("Payments", "#/app/payments", "parent", "Parent can never set status to Payé"),
    ("Live tracking", "#/app/live", "parent", "Route map + time slider"),
    ("Library", "#/app/library", "schooladmin, admin, teacher", "FEATURE FLAG OFF by default"),
    ("Transport", "#/app/transport", "schooladmin, admin, parent", "FEATURE FLAG OFF by default"),
    ("Events", "#/app/events", "6 roles", "Month calendar"),
    ("Social", "#/app/social", "6 roles", "LARGEST PAGE — RSVP / quorum / approval state machine"),
    ("Security", "#/app/security", "security, schooladmin, admin", "5 tabs"),
    ("Incidents", "#/app/incidents", "supervisor, security, admin, schooladmin", ""),
    ("Accidents", "#/app/accidents", "teacher, supervisor, admin, schooladmin, parent", "Two-eyes rule"),
    ("Behavior", "#/app/behavior", "teacher, admin, schooladmin, parent", "No ranking allowed"),
    ("Gallery", "#/app/gallery", "teacher, admin, schooladmin, parent", "Photo privacy per parent"),
    ("Canteen", "#/app/canteen", "teacher, admin, schooladmin, parent", "Allergen alerts"),
    ("Documents", "#/app/documents", "admin, schooladmin", "Gapless numbering, append-only"),
    ("Budget", "#/app/budget", "schooladmin, admin", ""),
    ("Inventory", "#/app/inventory", "schooladmin, admin", "Stock never below zero"),
    ("Recruit", "#/app/recruit", "schooladmin, admin", "No stage skipping"),
    ("Admissions", "#/app/admissions", "schooladmin, admin", "Capacity decides / waitlist"),
    ("Interop (OneRoster)", "#/app/interop", "schooladmin, admin", "CSV v1.2 export"),
    ("Requests", "#/app/requests", "teacher, admin, schooladmin", "12 types, 2-level approval"),
    ("Messages", "#/app/messages", "all 7", ""),
    ("Notices", "#/app/notices", "all 7", ""),
    ("Notifications", "#/app/notifications", "all 7", ""),
    ("Cross-cutting: AppShell", "(every page)", "all 7", "Sidebar, Ctrl+K palette, bell, user menu, FR/AR"),
    ("Cross-cutting: Arabic RTL", "(every page)", "all 7", "dir=rtl, Tajawal font, no clipped layout"),
    ("Cross-cutting: Mobile 390px", "(every page)", "all 7", "No horizontal overflow"),
]

MOD_HEADERS = ["#", "Module / Feature", "Route (URL)", "Roles that can access", "Notes / risk",
               "Owner", "Status", "Elements found", "Tested", "Passed", "Failed",
               "Coverage %", "Last tested", "Comments"]
MOD_WIDTHS = [5, 26, 24, 34, 42, 11, 14, 13, 9, 9, 8, 11, 12, 26]

md = sheet("Modules", MOD_HEADERS, MOD_WIDTHS,
           notes="MODULE INVENTORY — 51 testable surfaces.  Work through them one at a time.  "
                 "Coverage % fills itself from the Tested / Elements found columns.",
           freeze="C3", tab="2F8050")

for i, (mod, route, roles, note) in enumerate(MODULES, start=3):
    md.cell(row=i, column=1, value=i - 2)
    md.cell(row=i, column=2, value=mod).font = Font(size=10, bold=True)
    md.cell(row=i, column=3, value=route)
    md.cell(row=i, column=4, value=roles)
    md.cell(row=i, column=5, value=note)
    md.cell(row=i, column=7, value="Not started")
    # coverage % = tested / elements found
    md.cell(row=i, column=12,
            value=f'=IF(N(H{i})=0,"",J{i}/H{i})').number_format = "0%"
    for c in range(1, len(MOD_HEADERS) + 1):
        cell = md.cell(row=i, column=c)
        cell.border = BOX
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if cell.font.size is None or not cell.font.bold:
            cell.font = Font(size=10)

last_mod = 2 + len(MODULES)
dropdown(md, named("ModuleStatus"), "G", last=last_mod)
colour_rule(md, "G", {"Complete": (GREEN, GREEN_T), "Blocked": (RED, RED_T),
                      "In progress": (AMBER, AMBER_T), "Not started": (GREY, GREY_T)},
            last=last_mod)


# ================================================================ 6. ROLES & LOGINS
RL_HEADERS = ["Role (code)", "Label in app", "Demo e-mail", "Password", "What to check with this role", "Notes"]
RL_WIDTHS = [15, 18, 26, 14, 56, 34]
rl = sheet("Roles & Logins", RL_HEADERS, RL_WIDTHS,
           notes="THE 7 DEMO ACCOUNTS — always test with these. Never invent accounts: results must be reproducible.",
           tab="C2410C")

ROLES = [
    ("owner", "Plateforme", "owner@kogia.tn", "owner",
     "SaaS console: create a school, suspend it, reset demo data.", "Kogia Group level, above the school"),
    ("schooladmin", "Direction", "direction@alnour.tn", "admin",
     "Everything. Settings, accounts, modules, finance, promotion.", "The last active Direction account cannot be disabled"),
    ("admin", "Administration", "admin@alnour.tn", "office",
     "Office work: invoices, HR, admissions, documents.", "No access to Settings or Accounts"),
    ("teacher", "Enseignant", "enseignant@alnour.tn", "teacher",
     "Evaluate, attendance, behavior, gallery, requests.", "Must NOT receive payroll or invoice data"),
    ("teacher (crèche)", "Enseignant", "creche@alnour.tn", "teacher",
     "Day journal, meals, nap, diaper — early-years modules only.", "Level gating: must not see Timetable"),
    ("supervisor", "Surveillant", "surveillant@alnour.tn", "super",
     "Incidents, attendance, accidents.", "Restricted read scope"),
    ("security", "Sécurité", "securite@alnour.tn", "secu",
     "Visitors, rounds, log book, instructions.", "Must NOT see health, evaluations, behavior, photos"),
    ("parent", "Parent", "parent@alnour.tn", "parent",
     "Own child only: journal, payments, live, gallery, accidents.", "p1 has TWO children — use it to test the child selector"),
    ("parent 2-5", "Parent", "parent2..5@alnour.tn", "parent",
     "Privacy: must never see another family's child.", "Use to prove data isolation"),
]
for i, row in enumerate(ROLES, start=3):
    for j, v in enumerate(row, start=1):
        c = rl.cell(row=i, column=j, value=v)
        c.font = Font(size=10, bold=(j == 1))
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = BOX
    rl.row_dimensions[i].height = 30


# ================================================================ 7. TEST CYCLES
TC_HEADERS = ["Cycle #", "Name / release", "Build / Commit", "Start date", "End date", "Tester(s)",
              "Scope tested", "Cases run", "Passed", "Failed", "Blocked", "Pass rate",
              "Open Critical", "Open Major", "Go / No-Go", "Decided by", "Notes"]
TC_WIDTHS = [8, 22, 15, 11, 11, 14, 40, 10, 9, 8, 9, 10, 12, 11, 16, 13, 34]
tc = sheet("Test Cycles", TC_HEADERS, TC_WIDTHS,
           notes="TEST CYCLES — one row per round of testing. The Go / No-Go decision is recorded here, with its evidence.",
           tab="4361D0")
body_style(tc, len(TC_HEADERS), last=60)
for i in range(3, 61):
    tc.cell(row=i, column=12, value=f'=IF(N(H{i})=0,"",I{i}/H{i})').number_format = "0%"
dropdown(tc, named("Decision"), "O", last=60)
colour_rule(tc, "O", {"GO": (GREEN, GREEN_T), "NO-GO": (RED, RED_T),
                      "GO with conditions": (AMBER, AMBER_T),
                      "Pending": (GREY, GREY_T)}, last=60)


# ================================================================ 8. DASHBOARD
db = wb.create_sheet("Dashboard", 1)
db.sheet_properties.tabColor = NAVY
for col, w in zip("ABCDEF", [4, 34, 16, 16, 16, 44]):
    db.column_dimensions[col].width = w

db.merge_cells("B2:F2")
h = db["B2"]; h.value = "COREON EDU — QA DASHBOARD"
h.font = Font(bold=True, size=16, color="FFFFFF")
h.fill = PatternFill("solid", fgColor=NAVY)
h.alignment = Alignment(vertical="center", indent=1)
db.row_dimensions[2].height = 34

db["B3"] = "Read-only — every number is calculated from the other sheets. Do not type in column C."
db["B3"].font = Font(size=10, italic=True, color="7A8194")

def block(row, title):
    db.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    c = db.cell(row=row, column=2, value=title)
    c.font = Font(bold=True, size=11, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=PURPLE)
    c.alignment = Alignment(vertical="center", indent=1)
    db.row_dimensions[row].height = 22

def kpi(row, label, formula, fmt="0", hint=""):
    db.cell(row=row, column=2, value=label).font = Font(size=10, bold=True)
    c = db.cell(row=row, column=3, value=formula)
    c.font = Font(size=11, bold=True)
    c.alignment = Alignment(horizontal="center")
    c.fill = PatternFill("solid", fgColor=LIGHT)
    c.border = BOX
    c.number_format = fmt
    db.cell(row=row, column=6, value=hint).font = Font(size=9, color="7A8194")

T = "'Test Log'"
D = "'Defect Log'"
M = "Modules"

block(5, "TEST EXECUTION")
kpi(6,  "Test cases logged",  f"=COUNTA({T}!A3:A600)", hint="Rows filled in the Test Log")
kpi(7,  "Executed",           f'=COUNTIFS({T}!T3:T600,"PASS")+COUNTIFS({T}!T3:T600,"FAIL")+COUNTIFS({T}!T3:T600,"BLOCKED")', hint="Pass + Fail + Blocked")
kpi(8,  "Passed",             f'=COUNTIF({T}!T3:T600,"PASS")')
kpi(9,  "Failed",             f'=COUNTIF({T}!T3:T600,"FAIL")')
kpi(10, "Blocked",            f'=COUNTIF({T}!T3:T600,"BLOCKED")', hint="Could not be tested — not the same as a failure")
kpi(11, "Not run",            f'=COUNTIF({T}!T3:T600,"NOT RUN")')
kpi(12, "Pass rate",          f'=IF(C7=0,"",C8/C7)', fmt="0.0%", hint="Passed / executed")
kpi(13, "Execution progress", f'=IF(C6=0,"",C7/C6)', fmt="0.0%", hint="Executed / logged")

block(15, "DEFECTS")
kpi(16, "Total defects",  f"=COUNTA({D}!A3:A400)")
kpi(17, "Open",           f'=COUNTA({D}!A3:A400)-COUNTIFS({D}!U3:U400,"CLOSED")-COUNTIFS({D}!U3:U400,"WON\'T FIX")-COUNTIFS({D}!U3:U400,"DUPLICATE")-COUNTIFS({D}!U3:U400,"NOT A BUG")')
kpi(18, "Open Critical",  f'=COUNTIFS({D}!O3:O400,"Critical",{D}!U3:U400,"<>CLOSED")', hint="MUST be 0 before any release")
kpi(19, "Open Major",     f'=COUNTIFS({D}!O3:O400,"Major",{D}!U3:U400,"<>CLOSED")')
kpi(20, "Open Minor",     f'=COUNTIFS({D}!O3:O400,"Minor",{D}!U3:U400,"<>CLOSED")')
kpi(21, "Open Trivial",   f'=COUNTIFS({D}!O3:O400,"Trivial",{D}!U3:U400,"<>CLOSED")')
kpi(22, "P1 outstanding", f'=COUNTIFS({D}!P3:P400,"P1 - Immediate",{D}!U3:U400,"<>CLOSED")')
kpi(23, "Awaiting retest",f'=COUNTIF({D}!U3:U400,"READY TO RETEST")')
kpi(24, "Reopened",       f'=COUNTIF({D}!U3:U400,"REOPENED")', hint="High reopen count = fixes are not verified")

block(26, "MODULE COVERAGE")
kpi(27, "Modules total",       f"=COUNTA({M}!B3:B60)")
kpi(28, "Complete",            f'=COUNTIF({M}!G3:G60,"Complete")')
kpi(29, "In progress",         f'=COUNTIF({M}!G3:G60,"In progress")')
kpi(30, "Not started",         f'=COUNTIF({M}!G3:G60,"Not started")')
kpi(31, "Blocked",             f'=COUNTIF({M}!G3:G60,"Blocked")')
kpi(32, "Modules complete %",  f'=IF(C27=0,"",C28/C27)', fmt="0.0%")
kpi(33, "Elements identified", f"=SUM({M}!H3:H60)")
kpi(34, "Elements tested",     f"=SUM({M}!I3:I60)")
kpi(35, "Element coverage %",  f'=IF(C33=0,"",C34/C33)', fmt="0.0%", hint="The real 'have I clicked everything' number")

block(37, "EXIT CRITERIA — may we release?")
crit = [
    ("No test case left Not Run",           f'=IF(C11=0,"MET","NOT MET")'),
    ("Zero open Critical defects",          f'=IF(C18=0,"MET","NOT MET")'),
    ("Zero open P1 defects",                f'=IF(C22=0,"MET","NOT MET")'),
    ("Pass rate at or above 95%",           f'=IF(N(C12)>=0.95,"MET","NOT MET")'),
    ("All modules Complete or Out of scope",f'=IF(C30+C29+C31=0,"MET","NOT MET")'),
    ("Every fixed defect retested",         f'=IF(C23=0,"MET","NOT MET")'),
]
for i, (label, f) in enumerate(crit, start=38):
    db.cell(row=i, column=2, value=label).font = Font(size=10, bold=True)
    c = db.cell(row=i, column=3, value=f)
    c.font = Font(size=10, bold=True)
    c.alignment = Alignment(horizontal="center")
    c.border = BOX
colour_rule(db, "C", {"MET": (GREEN, GREEN_T), "NOT MET": (RED, RED_T)}, first=38, last=43)

db.merge_cells("B45:F45")
v = db["B45"]
v.value = ("VERDICT:  release only when every line above says MET.  "
           "If you override it, write down who decided and why — in Test Cycles.")
v.font = Font(bold=True, size=10, color="FFFFFF")
v.fill = PatternFill("solid", fgColor=NAVY)
v.alignment = Alignment(vertical="center", indent=1)
db.row_dimensions[45].height = 22

db.sheet_view.showGridLines = False
ws.sheet_view.showGridLines = False

# ================================================================ 9. RAW NOTES
import json, os

RAW = "/tmp/raw_notes.json"
raw_notes = json.load(open(RAW, encoding="utf-8")) if os.path.exists(RAW) else []

if raw_notes:
    rn = sheet("Raw Notes (original)", ["Note #", "Issue as originally written", "Location as originally written"],
               [8, 110, 40],
               notes="YOUR ORIGINAL NOTES — kept word for word, never edited. "
                     "Every row below is traced into the Defect Log or the Change Requests sheet.",
               tab=GREY_T)
    for i, nt in enumerate(raw_notes, start=3):
        rn.cell(row=i, column=1, value=nt["n"]).font = Font(size=10, bold=True)
        rn.cell(row=i, column=2, value=nt["issue"]).font = Font(size=10)
        rn.cell(row=i, column=3, value=nt["loc"]).font = Font(size=10)
        for c in range(1, 4):
            cell = rn.cell(row=i, column=c)
            cell.border = BOX
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        rn.row_dimensions[i].height = min(150, 15 + 11 * (len(nt["issue"]) // 110))


# ================================================================ 10. DEFECTS FOUND
# Real defects triaged out of the raw notes: things that are BROKEN in what already exists.
DEFECTS = [
    ("DEF-001", 11, "Landing & Login", "Login", "#/login", "Whole page", "all",
     "The login page does not fit the browser window",
     "1. Open #/login on a desktop browser\n2. Observe the page against the window edges\n3. Resize the window",
     "The page fits the viewport at every width, with no horizontal scrollbar and no clipped content.",
     "The layout does not fit the browser window; the page reads as broken.",
     "Major", "P2 - High", "Always",
     "Reported as 'page design is terrible and size page not fit the browser page'. "
     "Split from the design rework, which is tracked as CR-011."),

    ("DEF-002", 15, "Dashboard", "Dashboard", "#/app", "Several tiles / elements", "owner",
     "Several dashboard elements are not clickable",
     "1. Log in as owner@kogia.tn\n2. Open #/app\n3. Click each tile, figure and card in turn\n4. Note every one that does nothing",
     "Every tile and figure either opens a modal, navigates somewhere, or is visibly non-interactive by design.",
     "Some elements look clickable but do nothing when clicked.",
     "Major", "P2 - High", "Always",
     "List the exact elements in the Test Log before fixing. "
     "e2e/outil.audit-clics.mjs already detects dead buttons — run it first to get the full list."),

    ("DEF-003", 6, "Landing & Login", "Landing", "#/", "Anchor nav: Deux mondes / Modules / Vos données / Tarifs / FAQ", "public",
     "Landing anchor navigation gives no feedback that anything happened",
     "1. Open the landing page\n2. Click 'Modules' in the nav\n3. Observe the nav and the URL",
     "The clicked item is visibly marked as active and the user can tell the view changed.",
     "The page only scrolls; nothing in the nav changes, so the click feels like it did nothing.",
     "Minor", "P3 - Medium", "Always",
     "This is the defect half of note 6. Turning the one-pager into real routed pages is a separate "
     "product decision, tracked as CR-006."),

    ("DEF-004", 19, "Dashboard", "Dashboard", "#/app", "'Coreon Intelligence' heading block", "admin",
     "'Coreon Intelligence — ce que vos données disent cette semaine' is not aligned on one line",
     "1. Log in as admin@alnour.tn\n2. Open #/app\n3. Look at the Coreon Intelligence heading",
     "The title and its subtitle sit on a single consistent baseline.",
     "There is a small vertical drop between the two parts; they do not line up.",
     "Trivial", "P3 - Medium", "Always",
     "Check at several widths — likely a flex alignment issue, so it may only show at some breakpoints."),
]

for i, d in enumerate(DEFECTS, start=3):
    (did, note_n, module, screen, route, element, role, summary, steps,
     expected, actual, sev, prio, repro, comment) = d
    vals = [did, "=TODAY()", "Othman", f"(raw note {note_n})", "50d4c48",
            module, screen, route, element, role, summary, steps, expected, actual,
            sev, prio, repro, "edu.kogiagroup.com", "Desktop Chrome", "",
            "NEW", "Othman", "", "", "", comment]
    for j, v in enumerate(vals, start=1):
        c = dl.cell(row=i, column=j, value=v)
        c.font = Font(size=10)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = BOX
    dl.row_dimensions[i].height = 92


# ================================================================ 11. CHANGE REQUESTS
CR_HEADERS = [
    "CR ID", "Date raised", "Raised by", "Raw note #", "Type", "Epic / Theme",
    "Title", "What was asked (your words, tidied)", "Why it matters",
    "Area / Module", "Route (URL)", "Roles affected",
    "Priority", "Effort", "Status", "Target release",
    "Depends on", "Acceptance criteria (how we know it is done)", "Linked defect", "Comments",
]
CR_WIDTHS = [10, 12, 11, 10, 15, 22, 34, 52, 40, 20, 26, 16, 15, 9, 14, 14, 12, 56, 12, 26]

cr = sheet("Change Requests", CR_HEADERS, CR_WIDTHS,
           notes="CHANGE REQUEST LOG — things to ADD or CHANGE (the product works, but it should be different). "
                 "A defect is 'it does not do what it should'. A change request is 'what it should do must change'.",
           freeze="G3", tab="4361D0")

CRS = [
    ("CR-001", 1, "Content", "Content & tone",
     "Remove the em dash (—) from all product copy",
     "Remove the '—' character everywhere it appears in product content.",
     "Consistent typography; the em dash reads as machine-written and is awkward in French and Arabic.",
     "All modules", "Global", "all", "P3 - Medium", "S",
     "Grep the whole repo (app, core, mobile, i18n) for '—'. Zero occurrences remain in user-facing strings. "
     "Both fr and ar locale files verified. No layout breaks from the replacement punctuation.", ""),

    ("CR-002", 2, "Content", "Landing & login",
     "Clarify or rename the 'Deux mondes' section",
     "The section title 'Deux mondes' is not self-explanatory; a visitor cannot tell what it means.",
     "The landing page has about five seconds to be understood. An unclear heading loses the visitor.",
     "Landing", "https://edu.kogiagroup.com/", "public", "P2 - High", "S",
     "A reader who has never seen Coreon understands what the section offers from the heading alone, "
     "without scrolling. Tested on three people who do not know the product.", ""),

    ("CR-003", 3, "Enhancement", "Internationalisation",
     "Show pricing in EUR instead of TND",
     "Convert the pricing cards to euro.",
     "Required for the European positioning; TND signals a local-only product.",
     "Landing / Settings", "https://edu.kogiagroup.com/", "public, schooladmin", "P2 - High", "M",
     "Pricing displays in EUR with correct formatting. Currency is driven by the settings currency field, "
     "not hard-coded. Existing TND schools are unaffected.", ""),

    ("CR-004", 4, "Enhancement", "Internationalisation",
     "Remove Tunisia-specific content from the product",
     "Eliminate the Tunisian framing from all product content and elements.",
     "Positions the product for an international market rather than one country.",
     "All modules", "Global", "all", "P2 - High", "L",
     "No Tunisian governorate list, CIN-only identity, TND-only currency, Tunisia-specific request types "
     "or '© 2026 Tunisie' footer is forced on a new school. Each becomes a configurable locale pack. "
     "Tunisia remains available as one option, not the default assumption.", ""),

    ("CR-005", 5, "Epic", "Internationalisation",
     "Reposition Coreon EDU for the European market",
     "Redirect the product to be international, European-facing.",
     "Strategic direction for the whole product; drives CR-003, CR-004 and the compliance work.",
     "All modules", "Global", "all", "P1 - Immediate", "XL",
     "A written positioning decision exists: target countries, languages, currency, and the compliance "
     "regime (GDPR rather than loi 2004-63). Every child CR references it. Landing copy, pricing, "
     "identity fields and legal text all follow from it.", ""),

    ("CR-006", 6, "Enhancement", "Landing & login",
     "Replace the one-page anchor nav with real pages",
     "Clicking Deux mondes / Modules / Vos données / Tarifs / FAQ only scrolls the same page. "
     "It should feel like navigating to a different place.",
     "Anchor-only navigation feels unresponsive and gives each topic no shareable URL and no SEO value.",
     "Landing", "https://edu.kogiagroup.com/", "public", "P2 - High", "M",
     "Each nav item leads to its own routed page with its own URL that can be shared and bookmarked. "
     "The browser back button works. The active item is marked in the nav.", "DEF-003"),

    ("CR-007", 7, "Design", "Pre-enrolment",
     "Widen the pre-enrolment form layout",
     "The form is too narrow; use more of the horizontal space on left and right.",
     "A cramped public form is the first impression a parent gets of the school's software.",
     "Pre-enrolment", "#/inscription", "public", "P3 - Medium", "S",
     "The form uses a wider container on desktop, stays readable (line length under ~75 characters), "
     "and still shows zero horizontal overflow at 390 px.", ""),

    ("CR-008", 8, "New feature", "Pre-enrolment",
     "Extend the pre-enrolment form with full family and health details",
     "Add fields modelled on a standard preschool registration form: child first/last name, date of birth, "
     "gender, full home address; mother and father each with name, phone and address; health conditions "
     "(allergies, illness, disability) with a free-text description; anything else the teachers should know; "
     "preferred programme (days per week, before/after school care).",
     "The current form captures the minimum. A school cannot make an admission decision from it, and the "
     "data has to be re-collected by hand later.",
     "Pre-enrolment", "#/inscription", "public, schooladmin, admin", "P1 - Immediate", "L",
     "All new fields are captured, validated, persisted, and visible in the Admissions detail modal. "
     "Required vs optional is explicit. The health section is stored with the same protection as other "
     "health data. The form still submits successfully with only the required fields filled.", ""),

    ("CR-009", 9, "New feature", "Pre-enrolment",
     "Add terms, agreements and consent to the pre-enrolment form",
     "Select the applicable conditions from a standard registration form: fee payment commitment, "
     "notice period for withdrawal, immunisation record requirement, the school's right to decline, "
     "guardian signature and date, and data-protection consent.",
     "Without an agreed set of terms the pre-enrolment is not a commitment and gives the school no legal footing.",
     "Pre-enrolment", "#/inscription", "public, schooladmin", "P1 - Immediate", "M",
     "Terms are displayed before submission, consent is an explicit action (never pre-ticked), the submission "
     "is refused without it, and the accepted version and timestamp are stored with the application. "
     "Aligned with the compliance regime chosen in CR-005.", ""),

    ("CR-010", 10, "New feature", "Pre-enrolment",
     "Turn pre-enrolment into a multi-step wizard driven by the level",
     "The ten levels (Crèche, Pré-maternelle, Maternelle 1-2, 1ère to 6ème année) fall into categories with "
     "different rules and required information. The form should become several steps and ask only what that "
     "category needs.",
     "One flat form asking everything for everyone is long, confusing, and collects irrelevant data.",
     "Pre-enrolment", "#/inscription", "public", "P2 - High", "L",
     "Level categories are defined and documented. Choosing a level determines the following steps. "
     "Progress is visible, back and forward preserve entered data, and validation runs per step. "
     "Depends on CR-008 for the field set.", ""),

    ("CR-011", 11, "Design", "Landing & login",
     "Redesign the login page",
     "The login page design needs rework.",
     "The login screen is seen on every single visit by every user; it sets the perceived quality of the product.",
     "Login", "#/login", "all", "P2 - High", "M",
     "A redesign is agreed against the Kogia Harmony brand system, fits every viewport from 390 px upward, "
     "and keeps the seven demo role buttons available in demo mode only. "
     "The viewport defect DEF-001 is fixed as part of this.", "DEF-001"),

    ("CR-012", 12, "Content", "Landing & login",
     "Rewrite the login and landing marketing copy",
     "The current copy ('Pas un ERP scolaire de plus', 'L'école qu'on a envie d'ouvrir', the 121 / 96% / 8 "
     "figures, 'par Kogia Group © 2026 Tunisie') is not compelling.",
     "This is the first and most repeated text in the product; it carries the positioning.",
     "Landing / Login", "#/login", "public", "P2 - High", "M",
     "New copy is written after CR-005 fixes the positioning, drops the Tunisia reference, and states clearly "
     "what Coreon does and for whom. Demo statistics are either real or clearly labelled as a demonstration.", ""),

    ("CR-013", 13, "New feature", "Authentication",
     "Add a password recovery flow",
     "There is no 'forgot password' anywhere. Add one, handled through support@kogiagroup.com.",
     "Without recovery, a locked-out user cannot get back in without a developer. This blocks any real "
     "school using the product.",
     "Login", "#/login", "all", "P1 - Immediate", "M",
     "A 'Mot de passe oublié' link exists on the login page. Submitting an address sends a time-limited, "
     "single-use reset link and reveals nothing about whether the account exists. Tokens expire. "
     "The flow works in server mode; behaviour in demo mode is explicitly defined. "
     "Runs through the existing coreon-mail-worker.", ""),

    ("CR-014", 14, "New feature", "Owner console",
     "Add a technical corner to each school in the owner console",
     "As Kogia Group, be able to open a school and see its technical details: contact e-mails, domain, "
     "IP addresses, ports, configuration; and reach it directly to check what is happening. "
     "Start with whatever information exists today.",
     "You operate the platform for every school. Without this you cannot diagnose a customer problem "
     "without opening a terminal.",
     "Schools", "#/app/schools", "owner", "P2 - High", "L",
     "Each school row opens a technical panel showing the fields that exist today (name, plan, Direction "
     "contact, status, revision, last activity), with clearly marked empty placeholders for the fields "
     "not yet collected. Owner-only. Never exposes passwords or personal pupil data. "
     "A follow-up CR adds the missing infrastructure fields.", ""),

    ("CR-015", 16, "Design", "Shell & chrome",
     "Make the weather widget smaller",
     "The Météo corner is too prominent; reduce its size.",
     "It competes for attention with the working areas of the header.",
     "AppShell", "Global", "all", "P4 - Low", "XS",
     "The widget is visually secondary to the school name and the notification bell, still legible, "
     "and does not wrap the header at 390 px.", ""),

    ("CR-016", 17, "Research", "Role model",
     "Define what separates Direction from Administration",
     "The two roles feel like the same platform. The observed split is that Administration creates and "
     "receives requests and views evaluation dashboards, while Direction monitors and gives final approval. "
     "This needs to be researched and written down properly.",
     "Two roles that overlap confuse users, make permissions arbitrary, and block the departmental split in CR-019.",
     "Access control", "Global", "schooladmin, admin", "P1 - Immediate", "M",
     "A written role definition exists for both: what each one owns, decides, approves and cannot do. "
     "It is validated against how real schools are organised. ROUTE_ROLES and WRITE_ACL are updated to match, "
     "and the difference is visible in the product, not only in the code.", ""),

    ("CR-017", 18, "Epic", "Identity & numbering",
     "Give every record a formal, structured ID",
     "Nothing currently carries a proper identifier. Every entity — student, teacher, request, accident, "
     "registration — should have a structured ID following a real numbering convention, not a raw internal key.",
     "Structured IDs are what make an ERP auditable: they can be quoted on paper, searched, and traced.",
     "All modules", "Global", "all", "P1 - Immediate", "XL",
     "A documented numbering scheme exists (prefix, year, sequence, check digit) with one prefix per entity "
     "type. IDs are gapless, immutable once issued, never reused, and shown wherever the record appears. "
     "Existing demo data is migrated. Follows the pattern already proven in the Documents registry.", ""),

    ("CR-018", 18, "Epic", "Identity & numbering",
     "Add a person identity record separate from the entity ID",
     "Two different IDs are needed: the record ID from CR-017, and a real human identity for every person — "
     "national ID card or passport — chosen by the parent for a child and by the adult for themselves.",
     "One human can be a parent, an employee and an alumnus at once. Without a person record they become "
     "three unrelated rows, and no ERP can work that way.",
     "All modules", "Global", "all", "P1 - Immediate", "XL",
     "A person entity exists, holding identity document type, number, issuing country and expiry, with the "
     "document type selectable rather than fixed to CIN. One person links to many roles. Uniqueness is "
     "enforced per document type and country. Identity data is treated as sensitive and excluded from "
     "staff-facing blobs. Depends on CR-017 and on the compliance regime from CR-005.", ""),

    ("CR-019", 21, "Epic", "ERP architecture",
     "Split Administration into real departmental platforms",
     "One Administration platform holding every service is not how a school works — every school has "
     "departments, even when one person staffs them. Build genuine platforms for HR, Accounting and "
     "Administration.",
     "Departmental separation is what makes it an ERP rather than a collection of screens, and it is the "
     "precondition for real separation of duties.",
     "All modules", "Global", "schooladmin, admin", "P1 - Immediate", "XL",
     "Each department has its own workspace, its own role, its own permissions and its own dashboard. "
     "A person can hold more than one departmental role. Data crosses departments only through explicit, "
     "logged handovers. One person can still run all of them in a small school without extra friction. "
     "Depends on CR-016.", ""),

    ("CR-020", 22, "Research", "Role model",
     "Map the relationships between every role across every level",
     "Research how Direction, Administration, HR, Accounting, Supervisor, Security, Teacher and Parent relate "
     "to each other in a school, across all ten levels from Crèche to 6ème année.",
     "This is the foundation under CR-016, CR-019 and the permission model. Guessing it produces a product "
     "that no real school recognises.",
     "Access control", "Global", "all", "P1 - Immediate", "L",
     "A written model exists: who reports to whom, who approves what, which information flows between roles, "
     "and how this differs between early years and primary. Backed by real sources, not assumption. "
     "It becomes the specification the permission matrix is tested against.", ""),

    ("CR-021", 23, "Epic", "ERP architecture",
     "Reposition Coreon EDU as a genuine ERP",
     "The overall direction: move the product from its current state to a real ERP system.",
     "The parent decision that CR-016 through CR-019 all serve.",
     "All modules", "Global", "all", "P1 - Immediate", "XL",
     "A written definition of what 'real ERP' means for Coreon: departments, double-entry accounting, "
     "documented identity and numbering, audit trail, separation of duties, and period closing. "
     "A gap analysis lists what exists today against that definition, and every child CR traces back to it.", ""),

    ("CR-022", 24, "New feature", "Academic",
     "Let a teacher grade a whole class in one screen",
     "In Bulletins & passage, a teacher should be able to enter marks for the entire class across subjects "
     "(Mathématiques, Français, Arabe, Éveil scientifique, Anglais, each out of 20) rather than one pupil "
     "at a time.",
     "Entering marks pupil by pupil is the single most repetitive task a teacher does; it is where the "
     "product either saves time or wastes it.",
     "Academic", "#/app/academic", "teacher, schooladmin, admin", "P1 - Immediate", "L",
     "A grid shows pupils as rows and subjects as columns. Keyboard entry moves down the column. "
     "Marks are validated against the subject maximum. Partial entry can be saved and resumed. "
     "The subject list follows the level, and totals recalculate live.", ""),
]

for i, c_ in enumerate(CRS, start=3):
    (cid, note_n, ctype, epic, title, asked, why, area, route, roles,
     prio, effort, accept, linked) = c_
    vals = [cid, "=TODAY()", "Othman", note_n, ctype, epic, title, asked, why,
            area, route, roles, prio, effort, "NEW", "", "", accept, linked, ""]
    for j, v in enumerate(vals, start=1):
        cell = cr.cell(row=i, column=j, value=v)
        cell.font = Font(size=10, bold=(j == 7))
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = BOX
    cr.row_dimensions[i].height = 108

last_cr = 2 + len(CRS)
dropdown(cr, named("Priority"), "M", last=last_cr + 200)
dropdown(cr, named("CRType"), "E", last=last_cr + 200)
dropdown(cr, named("CRStatus"), "O", last=last_cr + 200)
dropdown(cr, named("Effort"), "N", last=last_cr + 200)
dropdown(cr, named("Role"), "L", last=last_cr + 200)

colour_rule(cr, "M", {"P1 - Immediate": (RED, RED_T), "P2 - High": (AMBER, AMBER_T)},
            last=last_cr + 200)
colour_rule(cr, "O", {"DONE": (GREEN, GREEN_T), "APPROVED": (GREEN, GREEN_T),
                      "IN PROGRESS": (AMBER, AMBER_T), "NEW": (GREY, GREY_T),
                      "TRIAGED": (GREY, GREY_T), "REJECTED": (RED, RED_T),
                      "DEFERRED": (GREY, GREY_T)}, last=last_cr + 200)
body_style(cr, len(CR_HEADERS), first=last_cr + 1, last=last_cr + 120)


# ---- extra dashboard block for change requests
block(47, "CHANGE REQUESTS")
C = "'Change Requests'"
kpi(48, "Total change requests", f"=COUNTA({C}!A3:A250)")
kpi(49, "P1 outstanding",        f'=COUNTIFS({C}!M3:M250,"P1 - Immediate",{C}!O3:O250,"<>DONE")')
kpi(50, "Epics",                 f'=COUNTIF({C}!E3:E250,"Epic")', hint="Big items that must be broken down further")
kpi(51, "Research items",        f'=COUNTIF({C}!E3:E250,"Research")', hint="Must be answered before their children can start")
kpi(52, "New features",          f'=COUNTIF({C}!E3:E250,"New feature")')
kpi(53, "Done",                  f'=COUNTIF({C}!O3:O250,"DONE")')
kpi(54, "Completion",            f'=IF(C48=0,"",C53/C48)', fmt="0.0%")


# order the tabs
order = ["READ ME", "Dashboard", "Test Log", "Defect Log", "Change Requests",
         "Modules", "Roles & Logins", "Test Cycles", "Raw Notes (original)", "Lists"]
wb._sheets = [wb[n] for n in order if n in wb.sheetnames]

OUT = "/mnt/c/Current LAB/Coreon-EDU_QA-Test-Workbook.xlsx"
wb.save(OUT)
print("saved:", OUT)
